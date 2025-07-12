import heapq
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from datetime import datetime, timedelta
import numpy as np
from .Structured_data import Operator, Product, Task, Machine
from collections import defaultdict
import pandas as pd

def boucle_principale(simulation, ordonnanceur,cout_global):
    """Exécute la simulation jusqu'à ce que toutes les tâches soient terminées ou que temps_max soit atteint."""
    nombre_de_taches=len(list(tache.id for tache in Task.instances))
    while len(simulation.termines)<nombre_de_taches:
        # 1. Vérifier si toutes les tâches sont terminées
        if all(prod.quantite_restante == 0 for prod in Product.instances):
            print("Toutes les tâches sont terminées !")
            break

        # 2. Récupérer le prochain événement à traiter
        if not simulation.evenements:
            print("Aucun événement à traiter.")
            break

        temps_evenement, evenement = heapq.heappop(simulation.evenements)
        simulation.temps_actuel = temps_evenement

        # 3. Traiter l'événement
        if evenement.type == "FIN_TACHE":
            print('fin de', evenement.tache.id)
            ordonnanceur.terminer_tache(evenement.tache, simulation)
            #metre à jour les performances
            temps_reel = evenement.tache.temps_standard +evenement.tache.temps_standard*(1-evenement.operateur.performancess [evenement.tache.id])
            # 2. Sous-performance (écart entre réel et standard)
            sous_performance =(-temps_reel + evenement.tache.temps_standard)*evenement.tache.quantite
            temps=evenement.temps
            evenement.operateur.mettre_a_jour_performance(evenement.tache.id,temps,simulation)
            cout_global.ajouter_cout(simulation, sous_performance, evenement.tache.cr, operateur=None, tache=None)
            print('le temps actuelle',  simulation.temps_actuel)
            # Chercher une nouvelle tâche pour l'opérateur
            tache_suivante = ordonnanceur.choisir_tache(
                operateur=evenement.operateur,
                simulation=simulation,
                tache_finie=evenement.tache
            )
            
            if tache_suivante:
                success = ordonnanceur.affecter_tache(
                    operateur=evenement.operateur,
                    tache=tache_suivante,
                    simulation=simulation
                )
                if not success:
                    print(f"Échec d'affectation pour {evenement.operateur.id}")

        elif evenement.type == "DEBUT_TACHE":

            print(f"Début de tâche {evenement.tache.id} à t={simulation.temps_actuel} ")

def export_gantt_to_excel(simulation, filename="gantt_data.xlsx"):
    """Exporte les données du diagramme Gantt vers un fichier Excel en utilisant des heures"""
    from datetime import datetime, timedelta
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    import os

    # 1. Récupération des données
    print("=== Collecte des données pour Excel ===")
    
    # Pour les opérateurs
    operator_data = []
    for op in Operator.instances:
        records = op.get_gantt_data()
        print(f"Opérateur {op.id}: {len(records)} tâches")
        for record in records:
            if record['Début'] is not None and record['Fin'] is not None:
                operator_data.append({
                    'Type': 'Opérateur',
                    'ID': f"Op{op.id}",
                    'Tâche': record['Tache'],
                    'Début (heures)': record['Début'] / 60,  # Convertir minutes en heures
                    'Fin (heures)': record['Fin'] / 60,      # Convertir minutes en heures
                    'Durée (heures)': (record['Fin'] - record['Début']) / 60,
                    'Machine': record.get('Machine', 'N/A')
                })
    
    # Pour les machines
    machine_data = []
    for mach in Machine.instances:
        records = mach.get_gantt_data()
        print(f"Machine {mach.id}: {len(records)} tâches")
        for record in records:
            if record['Début'] is not None and record['Fin'] is not None:
                machine_data.append({
                    'Type': 'Machine',
                    'ID': f"Mach{mach.id}",
                    'Tâche': record['Tache'],
                    'Début (heures)': record['Début'] / 60,  # Convertir minutes en heures
                    'Fin (heures)': record['Fin'] / 60,      # Convertir minutes en heures
                    'Durée (heures)': (record['Fin'] - record['Début']) / 60,
                    'Opérateur': record.get('Opérateur', 'N/A')
                })

    # 2. Création des DataFrames
    df_ops = pd.DataFrame(operator_data)
    df_machines = pd.DataFrame(machine_data)
    
    # 3. Création du fichier Excel avec mise en forme
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Feuille pour les opérateurs
        df_ops.to_excel(writer, sheet_name='Opérateurs', index=False)
        
        # Feuille pour les machines
        df_machines.to_excel(writer, sheet_name='Machines', index=False)
        
        # Formatage des colonnes
        workbook = writer.book
        
        for sheetname in ['Opérateurs', 'Machines']:
            worksheet = workbook[sheetname]
            
            # Ajustement de la largeur des colonnes
            for column in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                worksheet.column_dimensions[column].width = 18
            
            # Format numérique avec 2 décimales pour les heures
            for row in worksheet.iter_rows(min_row=2, max_col=5, max_row=worksheet.max_row):
                for cell in [row[3], row[4], row[5]]:  # Colonnes Début, Fin et Durée
                    cell.number_format = '0.00'

    print(f"Données sauvegardées dans {filename}")
    print(f"Emplacement: {os.path.abspath(filename)}")
 
def plot_individual_performance(operators):
    """
    Génère un graphique individuel pour chaque opérateur avec l'évolution de performance par tâche.
    """
    import matplotlib.pyplot as plt
    import os
    from collections import defaultdict

    os.makedirs("performance_operateurs", exist_ok=True)

    for op in operators:
        if not op.historique_performance:
            continue
        #print(op.id,op.derniere_execution)
       
       # Regrouper les performances par tâche
        evolution_par_tache = defaultdict(lambda: {'temps': [], 'perf': []})
        for tache_id, perf, temps in op.historique_performance:
            if tache_id in op.derniere_execution:
                evolution_par_tache[tache_id]['temps'].append(temps)
                evolution_par_tache[tache_id]['perf'].append(perf)
        # Tracer les courbes
        plt.figure(figsize=(14, 7))
        for tache_id, data in evolution_par_tache.items():
            plt.plot(data['temps'], data['perf'], label=tache_id)

        plt.title(f"Évolution des performances de l'opérateur {op.id}")
        plt.xlabel("Temps")
        plt.ylabel("Performance")
        plt.ylim(0.2, 1.2)
        plt.legend(loc='center left', bbox_to_anchor=(1, 0.5), fontsize='small')
        plt.grid(True)
        plt.tight_layout()
        plt.show()

def export_taches_affectees_excel(op, fichier_sortie="performance_taches_affectees.xlsx"):
    """
    Exporte les performances des tâches affectées vers un fichier Excel.
    
    Args:
        op (Operator): Objet opérateur contenant:
            - op.historique_performance: [[(tache, performance, temps), ...], ...]
            - op.taches_affectees: {ordre: tache_id}
        fichier_sortie (str): Chemin du fichier Excel de sortie.
    """
    # 1. Vérifier les données
    if not hasattr(op, 'taches_affectees') or not op.taches_affectees:
        raise ValueError("Aucune tâche affectée trouvée dans op.taches_affectees.")
    
    if not hasattr(op, 'historique_performance') or not op.historique_performance:
        raise ValueError("Aucune donnée de performance trouvée.")

    # 2. Filtrer les tâches affectées
    taches_affectees_ids = set(op.taches_affectees.values())
    donnees = defaultdict(list)

    for affectation in op.historique_performance:
        for record in affectation:
            try:
                tache, perf, temps = record
                if tache in taches_affectees_ids:
                    donnees["Tâche"].append(tache)
                    donnees["Performance"].append(perf)
                    donnees["Temps"].append(temps)
                    donnees["Affectation"].append(len(donnees["Temps"]) // len(taches_affectees_ids) + 1)
            except (ValueError, TypeError):
                continue

    # 3. Créer un DataFrame et exporter
    if not donnees:
        raise ValueError("Aucune donnée valide après filtrage.")

    df = pd.DataFrame(donnees)
    
    # Réorganiser les colonnes (optionnel)
    df = df[["Affectation", "Tâche", "Temps", "Performance"]]
    
    # Exporter vers Excel
    df.to_excel(fichier_sortie, index=False, sheet_name="Performance")
    print(f"Données exportées avec succès dans {fichier_sortie}")

def export_individual_performance_to_excel(operators, dossier="performance_excel"):
    """
    Exporte l'évolution des performances par tâche pour chaque opérateur dans un fichier Excel.
    """
    import os
    import pandas as pd
    from collections import defaultdict

    os.makedirs(dossier, exist_ok=True)

    for op in operators:
        if not op.historique_performance:
            continue

        # Regrouper les performances par tâche
        evolution_par_tache = defaultdict(list)

        for tache_id, perf, temps in op.historique_performance:
            evolution_par_tache['Tâche'].append(tache_id)
            evolution_par_tache['Temps'].append(temps)
            evolution_par_tache['Performance'].append(perf)

        # Créer DataFrame
        df = pd.DataFrame(evolution_par_tache)

        # Sauvegarder dans un fichier Excel
        nom_fichier = f"{dossier}/performance_{op.id}.xlsx"
        df.to_excel(nom_fichier, index=False)
        print(f"✅ Données exportées pour opérateur {op.id} : {nom_fichier}")

def get_data(operateurs,simulation,cout_global):
    evolution_par_tache_par_op={}
    operator_data = []
    for op in operateurs:
        if not op.historique_performance:
            continue
        #print(op.id,op.derniere_execution)
       
       # Regrouper les performances par tâche
        evolution_par_tache = defaultdict(lambda: {'temps': [], 'perf': []})
        for tache_id, perf, temps in op.historique_performance:
            if tache_id in op.derniere_execution:
                evolution_par_tache[tache_id]['temps'].append(temps)
                evolution_par_tache[tache_id]['perf'].append(perf)
        evolution_par_tache_par_op[op.id]=evolution_par_tache
        
        records = op.get_gantt_data()
        for record in records:
            if record['Début'] is not None and record['Fin'] is not None:
                operator_data.append({
                    'Type': 'Opérateur',
                    'ID': f"Op{op.id}",
                    'Tâche': record['Tache'],
                    'Début (heures)': record['Début'] / 60,  # Convertir minutes en heures
                    'Fin (heures)': record['Fin'] / 60,      # Convertir minutes en heures
                    'Durée (heures)': (record['Fin'] - record['Début']) / 60,
                    'Machine': record.get('Machine', 'N/A')
                })
    
    # Pour les machines
    machine_data = []
    for mach in Machine.instances:
        records = mach.get_gantt_data()
        print(f"Machine {mach.id}: {len(records)} tâches")
        for record in records:
            if record['Début'] is not None and record['Fin'] is not None:
                machine_data.append({
                    'Type': 'Machine',
                    'ID': f"Mach{mach.id}",
                    'Tâche': record['Tache'],
                    'Début (heures)': record['Début'] / 60,  # Convertir minutes en heures
                    'Fin (heures)': record['Fin'] / 60,      # Convertir minutes en heures
                    'Durée (heures)': (record['Fin'] - record['Début']) / 60,
                    'Opérateur': record.get('Opérateur', 'N/A')
                })

    data= {
        'gantt': {
            'operateurs': operator_data,
            'machines': machine_data
        },
        'performances': {
            'evolution_taches': evolution_par_tache_par_op,
            'makespan': simulation.makespan_actuel / 480,  # en heures
            'cout_total': cout_global.total_sous_performance
        }
    }

    return data